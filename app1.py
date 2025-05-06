import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer
from reportlab.lib.units import inch
import json
import os
from openpyxl import Workbook, load_workbook  # For handling Excel files
from openpyxl.utils.dataframe import dataframe_to_rows

# 1. Core Python Logic (Backend Brain)
def calculate_retirement_needs(age,
                           gender,
                           country,
                           income,
                           expenses,
                           assets_cars,
                           assets_land,
                           assets_others,
                           owns_house,
                           on_rent,
                           loans_debts_amount,
                           family_members_count,
                           dependents_parents,
                           dependents_spouse,
                           dependents_children,
                           dependent_health_problems_yn,
                           dependent_health_problems_details,
                           upcoming_big_goals,
                           health_issues_yn,
                           health_insurance_yn,
                           health_insurance_amount,
                           life_insurance_yn,
                           life_insurance_amount_total,
                           monthly_expenses,
                           pension_contributions_yn,
                           pension_contributions_amount,
                           expected_inflation_rate,
                           retirement_age_target,
                           only_source_of_income="N"):
    """
    Calculates retirement needs, corpus, and provides a verdict with suggestions.
    """

    # --- Constants and Assumptions ---
    REPLACEMENT_RATIO = 0.7
    POST_RETIREMENT_RETURN_RATE = 0.04
    YEARS_OF_RETIREMENT = 30
    INFLATION_ADJUSTMENT_RATE = 1 + expected_inflation_rate

    # --- Input Validation and Handling ---
    if any(val < 0 for val in [age, income, expenses, assets_cars, assets_land, assets_others, loans_debts_amount]):
        return {
            'corpus_needed': 0,
            'corpus_available': 0,
            'ready': False,
            'verdict': "Error: Please ensure all financial inputs are non-negative.",
            'suggestions': ["Please check your input values."]
        }
    if age >= retirement_age_target:
        return {
            'corpus_needed': 0,
            'corpus_available': 0,
            'ready': False,
            'verdict': "Error: Target retirement age should be greater than current age.",
            'suggestions': ["Please check your age and target retirement age."]
        }
    if owns_house not in ('Y', 'N'):
        return {
            'corpus_needed': 0,
            'corpus_available': 0,
            'ready': False,
            'verdict': "Error: Owns a House should be Y or N",
            'suggestions': ["Please check the Owns a House value."]
        }
    if on_rent not in ('Y', 'N'):
        return {
            'corpus_needed': 0,
            'corpus_available': 0,
            'ready': False,
            'verdict': "Error: On Rent should be Y or N",
            'suggestions': ["Please check the On Rent value."]
        }
    if dependent_health_problems_yn not in ('Y', 'N'):
        return {
            'corpus_needed': 0,
            'corpus_available': 0,
            'ready': False,
            'verdict': "Error: Dependent Health Problems should be Y or N",
            'suggestions': ["Please check the Dependent Health Problems value."]
        }
    if health_issues_yn not in ('Y', 'N'):
        return {
            'corpus_needed': 0,
            'corpus_available': 0,
            'ready': False,
            'verdict': "Error: Health Issues should be Y or N",
            'suggestions': ["Please check the Health Issues value."]
        }

    if health_insurance_yn not in ('Y', 'N'):
        return {
            'corpus_needed': 0,
            'corpus_available': 0,
            'ready': False,
            'verdict': "Error: Health Insurance should be Y or N",
            'suggestions': ["Please check the Health Insurance value."]
        }
    if life_insurance_yn not in ('Y', 'N'):
        return {
            'corpus_needed': 0,
            'corpus_available': 0,
            'ready': False,
            'verdict': "Error: Life Insurance should be Y or N",
            'suggestions': ["Please check the Life Insurance value."]
        }
    if pension_contributions_yn not in ('Y', 'N'):
        return {
            'corpus_needed': 0,
            'corpus_available': 0,
            'ready': False,
            'verdict': "Error: Pension Contributions should be Y or N",
            'suggestions': ["Please check the Pension Contributions value."]
        }
    if only_source_of_income not in ('Y', 'N'):
        return {
            'corpus_needed': 0,
            'corpus_available': 0,
            'ready': False,
            'verdict': "Error: Only Source of Income should be Y or N",
            'suggestions': ["Please check the Only Source of Income value."]
        }
    if dependents_spouse not in ('Y', 'N'):
        return {
            'corpus_needed': 0,
            'corpus_available': 0,
            'ready': False,
            'verdict': "Error: Dependents Spouse should be Y or N",
            'suggestions': ["Please check the Dependents Spouse value"]
        }

    # --- Calculations ---
    years_to_retirement = retirement_age_target - age
    inflation_factor = (INFLATION_ADJUSTMENT_RATE) ** years_to_retirement
    annual_expenses_at_retirement = (monthly_expenses * 12) * inflation_factor
    required_income_at_retirement = income * REPLACEMENT_RATIO * inflation_factor

    corpus_needed = required_income_at_retirement / POST_RETIREMENT_RETURN_RATE
    corpus_available = assets_cars + assets_land + assets_others - loans_debts_amount

    ready = corpus_available >= corpus_needed
    verdict = "Ready to Retire ✅" if ready else "Not Ready to Retire ❌"

    # --- Suggestions ---
    suggestions = []
    if not ready:
        shortfall = corpus_needed - corpus_available
        suggestions.append(f"Increase your retirement savings by ₹{shortfall:,.2f}.")
        suggestions.append(f"Consider increasing your savings rate.")

        if monthly_expenses * 12 > income * 0.8:
            suggestions.append("Reduce your expenses. Review your budget carefully.")

        if years_to_retirement < 10:
            suggestions.append(
                "You have a short time to prepare. Aggressively increase savings and consider delaying retirement.")
        elif years_to_retirement > 25:
            suggestions.append(
                "You have time to grow your retirement nest egg. Invest consistently and consider higher-growth investments.")

    if loans_debts_amount > 0:
        suggestions.append("Prioritize paying down high-interest debts.")

    if age < 40:
        suggestions.append("Start saving for retirement as early as possible to take advantage of compounding.")
    elif age > 50:
        suggestions.append("Focus on preserving your capital and reducing risk as you approach retirement.")

    if health_issues_yn == "Y":
        suggestions.append("Ensure you have adequate health insurance and consider long-term care planning.")

    if family_members_count > 2:
        suggestions.append("Consider the financial impact of supporting a larger family in retirement.")

    if dependents_children > 0:
        suggestions.append("Factor in future education and marriage expenses for your children.")

    if owns_house == "N" and on_rent == "Y":
        suggestions.append("Consider the impact of rent on your retirement expenses.  Evaluate if owning a home is feasible before retirement.")

    if pension_contributions_yn == "N":
        suggestions.append("Start contributing to a pension plan to build a retirement corpus.")

    if only_source_of_income == "Y":
        suggestions.append("Consider diversifying your income sources to reduce risk.")

    if dependents_spouse == "Y":
        suggestions.append("Ensure adequate financial planning for your spouse's needs in retirement.")

    if "Children's Education" in upcoming_big_goals:
        suggestions.append("Start a dedicated investment plan for your children's education.")

    return {
        'corpus_needed': corpus_needed,
        'corpus_available': corpus_available,
        'ready': ready,
        'verdict': verdict,
        'suggestions': suggestions,
    }


# 2. Build Interactive Web App with Streamlit
def main():
    """
    Main function to run the Streamlit app.
    """
    st.set_page_config(
        page_title="Retirement Readiness Checker",
        page_icon="💰",
        layout="wide",  # Use the wide layout
        initial_sidebar_state="auto",
    )

    # --- 3. Make it Look Nice ---
    # Add a title and logo
    col1, col2 = st.columns([1, 3])  # Adjust ratios as needed
    with col1:
        # Display an image from a URL
        st.image("cair.png",  # Replace with your image URL
                 width=250)  # Set the width of the image
    with col2:
        st.title("Retirement Readiness Checker")
    st.markdown("<h2 style='text-align: center;'>Assess your retirement readiness and get personalized recommendations.</h2>", unsafe_allow_html=True) # Centered

    # --- 1. Input Forms ---
    st.header("Your Information")
    age = st.number_input("Your Age", min_value=20, max_value=70, value=40)
    country = st.selectbox("Country",
                           ["USA", "India", "UK", "Canada", "Australia", "Germany", "Japan", "Brazil",
                            "South Africa", "France"], index=0)
    gender = st.selectbox("Gender", ["Male", "Female", "Other"], index=0)

    # Use a dictionary to map countries to their respective currencies and symbols
    currency_data = {
        "USA": {"symbol": "$", "code": "USD"},
        "India": {"symbol": "₹", "code": "INR"},
        "UK": {"symbol": "£", "code": "GBP"},
        "Canada": {"symbol": "$", "code": "CAD"},
        "Australia": {"symbol": "$", "code": "AUD"},
        "Germany": {"symbol": "€", "code": "EUR"},
        "Japan": {"symbol": "¥", "code": "JPY"},
        "Brazil": {"symbol": "R$", "code": "BRL"},
        "South Africa": {"symbol": "R", "code": "ZAR"},
        "France": {"symbol": "€", "code": "EUR"},
    }

    selected_currency = currency_data[country]["symbol"]
    selected_currency_code = currency_data[country]["code"]  # Get the currency code

    income = st.number_input(f"Net Annual Income ({selected_currency})", min_value=0, value=500000)
    only_source_of_income = st.selectbox("Only Source of Income", ["Y", "N"], index=0)
    expenses = st.number_input(f"Annual Expenses ({selected_currency})", min_value=0, value=300000)
    monthly_expenses = st.number_input(f"Monthly Expenses ({selected_currency})", min_value=0, value=25000)

    assets_cars = st.number_input(f"Assets (Cars) ({selected_currency})", min_value=0, value=100000)
    assets_land = st.number_input(f"Assets (Land) ({selected_currency})", min_value=0, value=200000)
    assets_others = st.number_input(f"Assets (Others) ({selected_currency})", min_value=0, value=300000)
    owns_house = st.selectbox("Owns a House (Y/N)", ["Y", "N"], index=0)
    on_rent = st.selectbox("On Rent (Y/N)", ["Y", "N"], index=0)
    loans_debts_amount = st.number_input(f"Loans/Debts Amount ({selected_currency})", min_value=0, value=100000)

    family_members_count = st.number_input("Family Members Count", min_value=1, value=4)
    dependents_parents = st.number_input("Dependents (Parents)", min_value=0, value=0)
    dependents_spouse = st.selectbox("Dependents (Spouse)", ["Y", "N"], index=0)
    dependents_children = st.number_input("Dependents (Children)", min_value=0, value=2)
    dependent_health_problems_yn = st.selectbox("Dependent Health Problems (Y/N)", ["Y", "N"], index=0)
    dependent_health_problems_details = st.text_input("Dependent Health Problems (Details)", "")
    upcoming_big_goals = st.text_input("Upcoming Big Goals", "Children's Education")

    health_issues_yn = st.selectbox("Any Health Issues (Y/N)", ["Y", "N"], index=0)
    health_insurance_yn = st.selectbox("Health Insurance (Y/N)", ["Y", "N"], index=0)
    health_insurance_amount = st.number_input(f"Health Insurance Amount (Monthly/Yearly) ({selected_currency})", min_value=0,
                                               value=10000)
    life_insurance_yn = st.selectbox("Life Insurance (Y/N)", ["Y", "N"], index=0)
    life_insurance_amount_total = st.number_input(f"Life Insurance Amount (Total) ({selected_currency})", min_value=0,
                                                   value=500000)

    pension_contributions_yn = st.selectbox("Pension Contributions (Y/N)", ["Y", "N"], index=0)
    pension_contributions_amount = st.number_input(
        f"Pension Contributions Amount (Monthly/Yearly) ({selected_currency})",
        min_value=0, value=5000)
    expected_inflation_rate = st.number_input("Expected Annual Inflation Rate (%)", min_value=0.0, max_value=10.0,
                                               value=6.0) / 100
    retirement_age_target = st.number_input("Target Retirement Age", min_value=50, max_value=75, value=60)
    # --- 2. Calculate and Display Results ---
    if st.button("Calculate Retirement Readiness"):
        results = calculate_retirement_needs(
            age,
            gender,
            country,
            income,
            expenses,  # Pass the expenses variable
            assets_cars,
            assets_land,
            assets_others,
            owns_house,
            on_rent,
            loans_debts_amount,
            family_members_count,
            dependents_parents, dependents_spouse,
            dependents_children, dependent_health_problems_yn,
            dependent_health_problems_details, upcoming_big_goals,
            health_issues_yn, health_insurance_yn, health_insurance_amount,
            life_insurance_yn, life_insurance_amount_total, monthly_expenses,
            pension_contributions_yn, pension_contributions_amount,
            expected_inflation_rate, retirement_age_target,
            only_source_of_income
        )

        st.header("Your Results")
        st.write(results['verdict'])

        col1, col2 = st.columns(2)
        with col1:
            st.subheader("Key Figures")
            st.metric(f"Corpus Needed ({selected_currency})", f"{results['corpus_needed']:,.2f}")
            st.metric(f"Corpus Available ({selected_currency})", f"{results['corpus_available']:,.2f}")

        with col2:
            fig, ax = plt.subplots()
            ax.bar([f'Corpus Available ({selected_currency})',
                    f'Corpus Needed ({selected_currency})'],
                   [results['corpus_available'], results['corpus_needed']],
                   color=['green' if results['ready'] else 'red', 'blue'])
            ax.set_ylabel(f"Amount ({selected_currency})")
            ax.set_title("Retirement Corpus Comparison")
            st.pyplot(fig)

        st.subheader("Suggestions")
        for suggestion in results['suggestions']:
            st.markdown(f"- {suggestion}")

        # --- (Optional) Download Button ---
        pdf_bytes = create_pdf_report(results, age,  gender, country, income,
                                      assets_cars, assets_land,
                                      assets_others, owns_house, on_rent, loans_debts_amount,
                                      family_members_count, dependents_parents, dependents_spouse,
                                      dependents_children, dependent_health_problems_yn,
                                      dependent_health_problems_details, upcoming_big_goals,
                                      health_issues_yn, health_insurance_yn, health_insurance_amount,
                                      life_insurance_yn, life_insurance_amount_total, monthly_expenses,
                                      pension_contributions_yn, pension_contributions_amount,
                                      expected_inflation_rate, retirement_age_target,
                                      country, gender, selected_currency)
        st.download_button(
            label="Download Retirement Plan (PDF)",
            data=pdf_bytes,
            file_name="retirement_plan.pdf",
            mime="application/pdf",
        )

        # --- 3. (Future Modification) Save User Data ---
        save_user_data(
            country, age, gender, income,
            assets_cars, assets_land, assets_others, owns_house, on_rent,
            loans_debts_amount, family_members_count, dependents_parents, dependents_spouse,
            dependents_children, dependent_health_problems_yn, dependent_health_problems_details, upcoming_big_goals,
            health_issues_yn, health_insurance_yn, health_insurance_amount,
            life_insurance_yn, life_insurance_amount_total, monthly_expenses,
            pension_contributions_yn, pension_contributions_amount,
            expected_inflation_rate, retirement_age_target, results, only_source_of_income
        )

    # Currency Conversion Section
    st.header("Currency Converter") # Moved to main area
    amount_to_convert = st.number_input(f"Amount in {selected_currency_code}", min_value=0.0, value=1000.0)
    target_currency_code = st.selectbox("Convert to",
                                       ["USD", "INR", "GBP", "CAD", "AUD", "EUR", "JPY", "BRL", "ZAR", "CNY"],
                                       index=0)
    if st.button("Convert"):
        converted_amount = convert_currency(amount_to_convert, selected_currency_code,
                                            target_currency_code)
        if converted_amount is not None:
            st.write(
                f"{amount_to_convert:.2f} {selected_currency_code} is equal to {converted_amount:.2f} {target_currency_code}")
        else:
            st.error("Currency conversion failed. Please check the currency codes.")


def convert_currency(amount, from_currency, to_currency):
    """
    A placeholder function for currency conversion.  In a real application, you would
    use a currency conversion API.
    """
    conversion_rates = {
        "USD": {"INR": 74.50, "EUR": 0.85, "GBP": 0.75, "CAD": 1.25, "AUD": 1.35, "JPY": 110.00,
                "BRL": 5.50, "ZAR": 15.00, "CNY": 6.45, "USD": 1.00},
        "INR": {"USD": 0.013, "EUR": 0.011, "GBP": 0.010, "CAD": 0.017, "AUD": 0.018, "JPY": 1.47,
                "BRL": 0.074, "ZAR": 0.20, "CNY": 0.086, "INR": 1.00},
        "EUR": {"USD": 1.18, "INR": 87.77, "GBP": 0.88, "CAD": 1.47, "AUD": 1.59, "JPY": 129.41,
                "BRL": 6.47, "ZAR": 17.65, "CNY": 7.61, "EUR": 1.00},
        "GBP": {"USD": 1.33, "INR": 98.80, "EUR": 1.14, "CAD": 1.66, "AUD": 1.79, "JPY": 146.67,
                "BRL": 7.29, "ZAR": 19.88, "CNY": 8.57, "GBP": 1.00},
        "CAD": {"USD": 0.80, "INR": 59.60, "EUR": 0.68, "GBP": 0.60, "AUD": 1.08, "JPY": 88.00,
                "BRL": 4.40, "ZAR": 12.00, "CNY": 5.16, "CAD": 1.00},
        "AUD": {"USD": 0.74, "INR": 55.19, "EUR": 0.63, "GBP": 0.56, "CAD": 0.93, "JPY": 81.48,
                "BRL": 4.07, "ZAR": 11.11, "CNY": 4.78, "AUD": 1.00},
        "JPY": {"USD": 0.0091, "INR": 0.68, "EUR": 0.0077, "GBP": 0.0068, "CAD": 0.0114,
                "AUD": 0.0123, "BRL": 0.050, "ZAR": 0.136, "CNY": 0.0586, "JPY": 1.00},
        "BRL": {"USD": 0.182, "INR": 13.55, "EUR": 0.155, "GBP": 0.137, "CAD": 0.227, "AUD": 0.246,
                "JPY": 20.00, "ZAR": 2.73, "CNY": 1.173, "BRL": 1.00},
        "ZAR": {"USD": 0.067, "INR": 4.97, "EUR": 0.057, "GBP": 0.050, "CAD": 0.083, "AUD": 0.090,
                "JPY": 7.33, "BRL": 0.367, "CNY": 0.430, "ZAR": 1.00},
        "CNY": {"USD": 0.155, "INR": 11.63, "EUR": 0.131, "GBP": 0.117, "CAD": 0.194, "AUD": 0.209,
                "JPY": 17.07, "BRL": 0.852, "ZAR": 2.326, "CNY": 1.00}
    }

    if from_currency in conversion_rates and to_currency in conversion_rates[from_currency]:
        rate = conversion_rates[from_currency][to_currency]
        return amount * rate
    else:
        return None


def create_pdf_report(results, age,  gender, country, income,
                      assets_cars, assets_land,
                      assets_others, owns_house, on_rent, loans_debts_amount,
                      family_members_count, dependents_parents, dependents_spouse,
                      dependents_children, dependent_health_problems_yn,
                      dependent_health_problems_details, upcoming_big_goals,
                      health_issues_yn, health_insurance_yn, health_insurance_amount,
                      life_insurance_yn, life_insurance_amount_total, monthly_expenses,
                      pension_contributions_yn, pension_contributions_amount,
                      expected_inflation_rate, retirement_age_target,
                      country_name, gender_name, currency_symbol):
    """
    Generates a PDF report of the retirement readiness assessment.

    Args:
        results (dict): The results dictionary from the calculate_retirement_needs function.
         age (int): Current age.
        income (float): Annual income.
        expenses (float): Annual expenses.
        assets_cars (float): Value of cars.
        assets_land (float): Value of land.
        assets_others (float): Value of other assets.
        owns_house (str): "Y" or "N".
        on_rent (str): "Y" or "N".
        loans_debts_amount (float): Total loans/debts.
        family_members_count (int): Number of family members.
        dependents_parents (int): Number of dependent parents.
        dependents_spouse (str): Y or N
        dependents_children (int): Number of dependent children.
        dependent_health_problems_yn (str): "Y" or "N".
        dependent_health_problems_details (str): Details of health problems.
        upcoming_big_goals (str): Upcoming financial goals.
        health_issues_yn (str): "Y" or "N".
        health_insurance_yn (str): "Y" or "N".
        health_insurance_amount (float): Health insurance amount.
        life_insurance_yn (str): "Y" or "N".
        life_insurance_amount_total (float): Life insurance amount.
        monthly_expenses (float): Monthly expenses.
        pension_contributions_yn (str): "Y" or "N".
        pension_contributions_amount (float): Pension contributions.
        expected_inflation_rate (float): Expected annual inflation rate.
        retirement_age_target (int): Target retirement age.
        country_name (str): Country Name
        gender_name (str): Gender Name
        currency_symbol (str): The currency symbol.
    Returns:
        bytes: The PDF report in bytes.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    Story = []

    # --- Title and Introduction ---
    Story.append(Paragraph("Retirement Readiness Report", styles['Title']))
    Story.append(Spacer(1, 0.2 * inch))
    Story.append(Paragraph("This report assesses your retirement readiness.", styles['Normal']))
    Story.append(Spacer(1, 0.2 * inch))

    # --- User Information ---
    Story.append(Paragraph("Your Information", styles['Heading1']))
    Story.append(Spacer(1, 0.1 * inch))
    user_info_text = f"Age: {age}, Gender: {gender_name}, Country: {country_name}, Net Annual Income: {currency_symbol}{income:,.2f},  Assets (Cars): {currency_symbol}{assets_cars:,.2f}, Assets (Land): {currency_symbol}{assets_land:,.2f}, Assets (Others): {currency_symbol}{assets_others:,.2f}, Owns a House: {owns_house}, On Rent: {on_rent}, Loans/Debts Amount: {currency_symbol}{loans_debts_amount:,.2f}, Family Members Count: {family_members_count}, Dependents (Parents): {dependents_parents}, Dependents (Spouse): {dependents_spouse}, Dependents (Children): {dependents_children}, Dependent Health Problems: {dependent_health_problems_yn}, Dependent Health Problems Details: {dependent_health_problems_details}, Upcoming Big Goals: {upcoming_big_goals}, Any Health Issues: {health_issues_yn}, Health Insurance: {health_insurance_yn}, Health Insurance Amount (Monthly/Yearly): {currency_symbol}{health_insurance_amount:,.2f}, Life Insurance: {life_insurance_yn}, Life Insurance Amount (Total): {currency_symbol}{life_insurance_amount_total:,.2f}, Monthly Expenses: {currency_symbol}{monthly_expenses:,.2f}, Pension Contributions: {pension_contributions_yn}, Pension Contributions Amount (Monthly/Yearly): {currency_symbol}{pension_contributions_amount:,.2f}, Expected Inflation Rate: {expected_inflation_rate * 100:.2f}%, Target Retirement Age: {retirement_age_target}"
    Story.append(Paragraph(user_info_text, styles['Normal']))
    Story.append(Spacer(1, 0.2 * inch))

    # --- Key Results ---
    Story.append(Paragraph("Key Results", styles['Heading1']))
    Story.append(Spacer(1, 0.1 * inch))
    Story.append(Paragraph(f"Verdict: {results['verdict']}", styles['Heading2']))
    Story.append(Paragraph(f"Corpus Needed: {currency_symbol}{results['corpus_needed']:,.2f}", styles['Normal']))
    Story.append(Paragraph(f"Corpus Available: {currency_symbol}{results['corpus_available']:,.2f}", styles['Normal']))
    Story.append(Spacer(1, 0.2 * inch))

    # --- Suggestions ---
    Story.append(Paragraph("Recommendations", styles['Heading1']))
    Story.append(Spacer(1, 0.1 * inch))
    for suggestion in results['suggestions']:
        Story.append(Paragraph(f"- {suggestion}", styles['Bullet']))
    Story.append(Spacer(1, 0.2 * inch))

    # --- Additional Details (Optional) ---
    Story.append(Paragraph("Assumptions and Methodology", styles['Heading2']))
    Story.append(Spacer(1, 0.1 * inch))
    Story.append(Paragraph(
        "The calculations in this report are based on standard financial planning principles. It's essential to consult with a qualified financial advisor for personalized advice.",
        styles['Normal']
    ))

    doc.build(Story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes



def save_user_data(country, age, gender, income,
                   assets_cars, assets_land, assets_others, owns_house, on_rent,
                   loans_debts_amount, family_members_count, dependents_parents, dependents_spouse,
                   dependents_children, dependent_health_problems_yn, dependent_health_problems_details, upcoming_big_goals,
                   health_issues_yn, health_insurance_yn, health_insurance_amount,
                   life_insurance_yn, life_insurance_amount_total, monthly_expenses,
                   pension_contributions_yn, pension_contributions_amount,
                   expected_inflation_rate, retirement_age_target, results, only_source_of_income):
    """
    Saves user input data and results to a JSON file.  This is a simplified
    example; ina production environment, you would use a database.

    Args:
        All the input parameters from the main function and results dictionary
    """
    # Create a unique filename (e.g.,based on timestamp)
    #  Use age, and a few other unique identifiers
    filename = f"user_data.xlsx"
    filepath = os.path.join("user_data", filename) #saves the file in user_data directory

    # Ensure the 'user_data' directory exists
    if not os.path.exists("user_data"):
        os.makedirs("user_data")

    # Load the workbook if it exists, create a new one if it doesn't
    wb = load_workbook(filepath) if os.path.exists(filepath) else Workbook()
    sheet_name = 'Sheet1'

    # Create the sheet if it doesn't exist
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)

    sheet = wb[sheet_name] # get the sheet
    

    # Get the last row.
    last_row = sheet.max_row
    # If the sheet is empty, write the headers
    if  sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
        sheet.append(["Country", "Age", "Gender", "Income", "Assets Cars", "Assets Land", "Assets Others", "Owns House",
                     "On Rent", "Loans/Debts Amount", "Family Members Count", "Dependents Parents",
                     "Dependents Spouse", "Dependents Children", "Dependent Health Problems YN",
                     "Dependent Health Problems Details", "Upcoming Big Goals", "Any Health Issues",
                     "Health Insurance", "Health Insurance Amount", "Life Insurance",
                     "Life Insurance Amount Total", "Monthly Expenses", "Pension Contributions",
                     "Pension Contributions Amount", "Expected Inflation Rate", "Target Retirement Age",
                     "Results Verdict", "Corpus Needed", "Corpus Available", "Ready",
                     "Only Source of Income"])  # only add headers if the sheet is newly created

    # Append the data to the sheet
    sheet.append([country, age, gender, income,
                    assets_cars, assets_land, assets_others, owns_house, on_rent,
                    loans_debts_amount, family_members_count, dependents_parents, dependents_spouse,
                    dependents_children, dependent_health_problems_yn, dependent_health_problems_details, upcoming_big_goals,
                    health_issues_yn, health_insurance_yn, health_insurance_amount,
                    life_insurance_yn, life_insurance_amount_total, monthly_expenses,
                    pension_contributions_yn, pension_contributions_amount,
                    expected_inflation_rate, retirement_age_target, results["verdict"], results["corpus_needed"], results["corpus_available"], results["ready"], only_source_of_income]) # added only_source_of_income

    try:
        wb.save(filepath)
        print(f"User data saved to {filepath}")
    except Exception as e:
        print(f"Error saving user data: {e}")



if __name__ == "__main__":
    main()
